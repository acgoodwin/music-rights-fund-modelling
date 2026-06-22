#!/usr/bin/env python3
"""
Music Royalty Fund Financial Model
Models a $150M music royalty fund with staged deployment, debt financing, and exit.
Output from perspective of 44% shareholder in UK company (which owns 33.33% of mgmt co).
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from scipy.optimize import newton
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================================
# PARAMETERS - ALL ADJUSTABLE
# ============================================================================

class ModelParams:
    # Fund & Deployment
    total_equity = 150_000_000  # $150M
    deployment_years = 3
    avg_deal_size = 4_000_000  # $4M (variable)
    purchase_multiple = 10  # 10x revenues
    annual_asset_return = 0.10  # 10% annual return
    return_delay_months = 12  # 12-month lag
    hold_period_years = 5

    # Deal Costs (as % of 10x revenue purchase price)
    deal_cost_data_work = 0.015  # 1.5%
    deal_cost_valuation = 0.010  # 1.0%
    deal_cost_legal = 0.020  # 2.0%
    deal_cost_broker = 0.030  # 3.0%

    # Management Company Structure
    uk_company_ownership = 0.44  # User's 44% shareholding
    uk_company_owns_mgmt = 0.3333  # UK company owns 33.33% of mgmt co
    mgmt_fee_rate = 0.02  # 2% of AUM
    mgmt_fee_on_reinvested = True  # Toggle: apply fee to reinvested returns
    mgmt_upside_split = 0.20  # Mgmt co gets 20% of upside, investor 80%

    # Reinvestment
    reinvestment_rate = 0.50  # % of returns reinvested (toggle between 0-1)

    # Debt Financing
    debt_amount = 100_000_000  # $100M
    debt_raise_timing = 2.5  # Mid-year 3 (2.5 years from start)
    debt_interest_rate = 0.07  # 7% (adjustable)
    debt_arrangement_fee = 0.015  # 1.5%
    debt_legal_fee = 0.005  # 0.5% (estimate for legal costs)
    debt_tenor_years = 7
    debt_amortization_type = "bullet"  # "bullet" or "straight_line"

    # Exit
    exit_multiple_low = 14  # 14x
    exit_multiple_high = 18  # 18x
    exit_multiple_on_purchase_price = True  # Toggle: True = purchase price, False = asset value
    exit_costs_pct = 0.02  # 2% of exit value (legal, transaction costs)


# ============================================================================
# MODEL LOGIC
# ============================================================================

def build_deployment_schedule(params):
    """Build year-by-year deployment schedule."""
    total_years = params.hold_period_years + 1
    years = list(range(total_years))

    # Distribute equity deployment over deployment_years
    deployment_schedule = {}
    annual_deployment = params.total_equity / params.deployment_years

    for year in years:
        if year < params.deployment_years:
            deployment_schedule[year] = annual_deployment
        else:
            deployment_schedule[year] = 0

    return deployment_schedule, years


def build_asset_schedule(params, deployment_schedule):
    """Build asset acquisition and return schedule."""
    assets = []
    asset_id = 0

    for year, deployment_amount in deployment_schedule.items():
        if deployment_amount > 0:
            num_deals = int(deployment_amount / params.avg_deal_size)

            for deal_idx in range(num_deals):
                purchase_price = params.avg_deal_size  # $4M deal at 10x revenues
                deal_costs = (
                    purchase_price * (
                        params.deal_cost_data_work +
                        params.deal_cost_valuation +
                        params.deal_cost_legal +
                        params.deal_cost_broker
                    )
                )
                total_invested = purchase_price + deal_costs

                assets.append({
                    'asset_id': asset_id,
                    'acquisition_year': year,
                    'purchase_price': purchase_price,
                    'deal_costs_data': purchase_price * params.deal_cost_data_work,
                    'deal_costs_valuation': purchase_price * params.deal_cost_valuation,
                    'deal_costs_legal': purchase_price * params.deal_cost_legal,
                    'deal_costs_broker': purchase_price * params.deal_cost_broker,
                    'total_deal_costs': deal_costs,
                    'total_invested': total_invested,
                    'annual_return_pct': params.annual_asset_return,
                })
                asset_id += 1

    return assets


def calculate_asset_returns(assets, params, total_years):
    """Calculate year-by-year returns from assets."""
    annual_returns = {}

    for year in range(total_years):
        year_return = 0
        for asset in assets:
            years_since_acq = year - asset['acquisition_year']
            # Apply 12-month lag: returns start when years_since_acq >= (lag in years)
            if years_since_acq >= (params.return_delay_months / 12):
                # Annual return on invested amount
                base_return = asset['total_invested'] * params.annual_asset_return
                year_return += base_return

        annual_returns[year] = year_return

    return annual_returns


def build_financial_model(params):
    """Build complete financial model."""
    deployment_schedule, years = build_deployment_schedule(params)
    assets = build_asset_schedule(params, deployment_schedule)
    asset_returns = calculate_asset_returns(assets, params, len(years))

    # Initialize model structure
    model_data = {}

    for year in years:
        model_data[year] = {
            'year': year,
            'new_deployment': deployment_schedule.get(year, 0),
            'asset_returns_gross': asset_returns.get(year, 0),
            'cumulative_aum': 0,
            'mgmt_fee': 0,
            'available_cash': 0,
            'reinvestment_amount': 0,
            'cash_distributions': 0,
            'debt_drawn': 0,
            'debt_interest': 0,
            'debt_principal_paid': 0,
            'debt_outstanding': 0,
        }

    # Calculate cumulative AUM, fees, and cashflows
    cumulative_aum = 0  # Total assets under management
    cumulative_reinvested = 0
    debt_outstanding = 0
    debt_raised_year = None
    debt_available_for_deployment = 0

    for year in years:
        # New deployment
        new_deployment = model_data[year]['new_deployment']

        # Update cumulative AUM (assets purchased)
        cumulative_aum += new_deployment

        # Calculate AUM for fee purposes (before new deployment fee impact)
        aum_for_fee = cumulative_aum
        if params.mgmt_fee_on_reinvested:
            aum_for_fee += cumulative_reinvested

        # Management fee (2% of AUM)
        mgmt_fee = aum_for_fee * params.mgmt_fee_rate
        model_data[year]['mgmt_fee'] = mgmt_fee
        model_data[year]['cumulative_aum'] = aum_for_fee

        # Available cash = new deployment + asset returns - mgmt fee - debt interest
        gross_cash = new_deployment + asset_returns.get(year, 0)

        # Debt raised (mid-year 3), but deployed starting year 4
        if year >= params.debt_raise_timing and debt_raised_year is None:
            debt_net = params.debt_amount * (1 - (params.debt_arrangement_fee + params.debt_legal_fee))
            model_data[year]['debt_drawn'] = debt_net
            debt_outstanding = params.debt_amount
            debt_raised_year = year
            debt_available_for_deployment = debt_net
            # Don't add to gross_cash yet; it's only available starting next year

        # Deploy debt starting in the year AFTER it was raised
        if debt_raised_year is not None and year > debt_raised_year and debt_available_for_deployment > 0:
            gross_cash += debt_available_for_deployment
            debt_available_for_deployment = 0  # Only deploy once

        # Debt servicing
        debt_interest = 0
        debt_principal = 0
        if debt_outstanding > 0:
            debt_interest = debt_outstanding * params.debt_interest_rate
            model_data[year]['debt_interest'] = debt_interest

            # Principal repayment schedule
            if params.debt_amortization_type == "straight_line":
                debt_principal = params.debt_amount / params.debt_tenor_years
            else:  # bullet
                debt_principal = 0 if year < (params.debt_raise_timing + params.debt_tenor_years - 1) else params.debt_amount

            model_data[year]['debt_principal_paid'] = debt_principal
            debt_outstanding -= debt_principal

        model_data[year]['debt_outstanding'] = debt_outstanding

        # Available cash after fees and debt service
        available_cash = gross_cash - mgmt_fee - debt_interest - debt_principal

        # Reinvestment vs distribution
        reinvestment = available_cash * params.reinvestment_rate
        distribution = available_cash * (1 - params.reinvestment_rate)

        cumulative_aum += reinvestment
        cumulative_reinvested += reinvestment

        model_data[year]['available_cash'] = available_cash
        model_data[year]['reinvestment_amount'] = reinvestment
        model_data[year]['cash_distributions'] = distribution

    return model_data, assets, cumulative_aum, cumulative_reinvested


def calculate_exit_value(model_data, assets, params, cumulative_deployed, cumulative_reinvested):
    """Calculate exit value and returns to shareholder."""
    exit_year = params.hold_period_years

    if params.exit_multiple_on_purchase_price:
        # Exit on original purchase price
        base_value = sum([a['purchase_price'] for a in assets])
        exit_value = base_value * params.exit_multiple_low  # Use low end for now
    else:
        # Exit on total invested (including deal costs)
        base_value = sum([a['total_invested'] for a in assets])
        exit_value = base_value * params.exit_multiple_low

    exit_costs = exit_value * params.exit_costs_pct
    net_exit_proceeds = exit_value - exit_costs

    return {
        'exit_year': exit_year,
        'exit_value': exit_value,
        'exit_costs': exit_costs,
        'net_proceeds': net_exit_proceeds,
    }


def calculate_irr(model_data, exit_data, params):
    """Calculate IRR for investor."""
    years = sorted(model_data.keys())

    # Build investor cashflows from perspective of equity investor (80% investor portion)
    investor_cashflows = []

    for year in years:
        data = model_data[year]

        # Investor invests the new deployment (80% of new equity)
        equity_invested = data['new_deployment']

        # Investor receives asset returns minus fees (mgmt fee goes to mgmt co)
        asset_returns = data['asset_returns_gross']
        mgmt_fee = data['mgmt_fee']
        debt_interest = data['debt_interest']
        debt_principal = data['debt_principal_paid']

        # Investor's net cashflow = returns - fees - debt service
        investor_flow = asset_returns - mgmt_fee - debt_interest - debt_principal

        # If year 0, subtract initial investment
        if year == 0:
            investor_flow -= equity_invested

        # At exit year, add exit proceeds (80% to investor, 20% to mgmt co)
        if year == exit_data['exit_year']:
            exit_proceeds_to_investor = exit_data['net_proceeds'] * 0.80
            investor_flow += exit_proceeds_to_investor

        investor_cashflows.append(investor_flow)

    # Calculate IRR using NPV approach
    def npv(rate, cashflows):
        return sum(cf / ((1 + rate) ** i) for i, cf in enumerate(cashflows))

    def calculate_rate(cashflows):
        try:
            # Try multiple initial guesses for robustness
            for guess in [0.05, 0.10, 0.15, 0.20, 0.30]:
                try:
                    rate = newton(lambda r: npv(r, cashflows), guess, maxiter=100)
                    if -0.5 < rate < 2.0:  # Sanity check: -50% to 200%
                        return rate * 100
                except:
                    continue
            return 0
        except:
            return 0

    investor_irr = calculate_rate(investor_cashflows)

    return {
        'investor_irr': investor_irr,
        'investor_cashflows': investor_cashflows,
    }


def calculate_shareholder_returns(model_data, exit_data, params):
    """Calculate returns from 44% shareholder perspective."""
    # Shareholder ownership: 44% of UK company, which owns 33.33% of mgmt co
    # Shareholder %: 44% × 33.33% = 14.67%

    # Shareholder receives:
    # 1. Management fee dividend income: 14.67% of mgmt fee
    # 2. Upside at exit: 44% × 33.33% × (20% mgmt upside)

    shareholder_pct_of_mgmt = params.uk_company_ownership * params.uk_company_owns_mgmt

    # Calculate fee dividend income (annual)
    cumulative_fee_income = 0
    for year, data in model_data.items():
        mgmt_fee = data['mgmt_fee']
        shareholder_fee_income = mgmt_fee * shareholder_pct_of_mgmt
        cumulative_fee_income += shareholder_fee_income

    # Calculate upside at exit
    # Mgmt co owns 20% of upside
    # UK company owns 33.33% of mgmt co
    # Shareholder owns 44% of UK company
    mgmt_co_upside_pct = params.mgmt_upside_split  # 20%
    uk_company_upside_pct = mgmt_co_upside_pct * params.uk_company_owns_mgmt  # 20% × 33.33%
    shareholder_upside_pct = uk_company_upside_pct * params.uk_company_ownership  # × 44%

    # Exit value (net of costs)
    exit_value_net = exit_data['net_proceeds']
    shareholder_upside_income = exit_value_net * shareholder_upside_pct

    return {
        'shareholder_pct_of_mgmt': shareholder_pct_of_mgmt,
        'cumulative_fee_income': cumulative_fee_income,
        'shareholder_upside_pct': shareholder_upside_pct,
        'exit_upside_income': shareholder_upside_income,
        'total_cash_to_shareholder': cumulative_fee_income + shareholder_upside_income,
    }


# ============================================================================
# OUTPUT
# ============================================================================

def generate_timeline_html(model_data, params):
    """Generate horizontal timeline view (years as columns)."""
    years = sorted(model_data.keys())
    nav_tabs = generate_nav_tabs('financial_model_timeline.html')

    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Financial Model — Timeline View</title>
        <style>
            body { font-family: 'Courier New', monospace; margin: 20px; background-color: #f5f5f5; }
            h1 { color: #333; }
            .container { max-width: 2000px; margin: 0 auto; background-color: white; padding: 20px; border-radius: 5px; }
            table { border-collapse: collapse; margin: 20px 0; background-color: white; overflow-x: auto; }
            th, td { border: 1px solid #ddd; padding: 8px 12px; text-align: right; font-size: 0.9em; }
            th { background-color: #4CAF50; color: white; min-width: 120px; }
            td.label { text-align: left; font-weight: bold; background-color: #f9f9f9; min-width: 180px; }
            .positive { color: green; font-weight: bold; }
            .negative { color: red; font-weight: bold; }
            .header { background-color: #4CAF50; color: white; font-weight: bold; }
            .section { margin-top: 30px; }
            .scroll-hint { color: #999; font-size: 0.9em; margin-bottom: 10px; }
        </style>
    </head>
    <body>
        <div class="container">
            """ + nav_tabs + """
            <h1>Financial Model — Timeline View (Horizontal)</h1>
            <p style="margin-bottom: 15px;">
                <a href="financial_model.xlsx" style="display: inline-block; padding: 10px 20px; background-color: #4CAF50; color: white; text-decoration: none; border-radius: 5px; font-weight: bold;">
                    ⬇️ Download Excel
                </a>
            </p>
            <p class="scroll-hint">💡 Scroll right to see all years side-by-side</p>
    """

    # Deployments & Asset Returns
    html += """
    <div class="section">
    <h2>Capital Flows</h2>
    <table>
        <tr class="header">
            <th>Metric</th>
    """
    for year in years:
        html += f"<th>Year {year}</th>"
    html += "</tr>"

    # Row: New Deployment
    html += "<tr><td class='label'>New Deployment</td>"
    for year in years:
        val = model_data[year]['new_deployment']
        html += f"<td class='number'>${val/1e6:.1f}M</td>"
    html += "</tr>"

    # Row: Asset Returns
    html += "<tr><td class='label'>Asset Returns (Gross)</td>"
    for year in years:
        val = model_data[year]['asset_returns_gross']
        css = 'positive' if val > 0 else ''
        html += f"<td class='{css}'>${val/1e6:.1f}M</td>"
    html += "</tr>"

    html += "</table>"

    # Fees & Interest
    html += """
    <h2>Fees & Servicing</h2>
    <table>
        <tr class="header">
            <th>Metric</th>
    """
    for year in years:
        html += f"<th>Year {year}</th>"
    html += "</tr>"

    # Row: Mgmt Fee
    html += "<tr><td class='label'>Management Fee (2% AUM)</td>"
    for year in years:
        val = model_data[year]['mgmt_fee']
        html += f"<td class='negative'>${val/1e6:.1f}M</td>"
    html += "</tr>"

    # Row: Debt Interest
    html += "<tr><td class='label'>Debt Interest</td>"
    for year in years:
        val = model_data[year]['debt_interest']
        css = 'negative' if val > 0 else ''
        html += f"<td class='{css}'>${val/1e6:.1f}M</td>"
    html += "</tr>"

    # Row: Debt Drawn
    html += "<tr><td class='label'>Debt Drawn</td>"
    for year in years:
        val = model_data[year]['debt_drawn']
        css = 'positive' if val > 0 else ''
        html += f"<td class='{css}'>${val/1e6:.1f}M</td>"
    html += "</tr>"

    html += "</table>"

    # Cashflow & Reinvestment
    html += """
    <h2>Reinvestment & Distribution</h2>
    <table>
        <tr class="header">
            <th>Metric</th>
    """
    for year in years:
        html += f"<th>Year {year}</th>"
    html += "</tr>"

    # Row: Available Cash
    html += "<tr><td class='label'>Available Cash (after fees)</td>"
    for year in years:
        val = model_data[year]['available_cash']
        css = 'positive' if val > 0 else 'negative'
        html += f"<td class='{css}'>${val/1e6:.1f}M</td>"
    html += "</tr>"

    # Row: Reinvestment
    html += "<tr><td class='label'>Reinvested ({:.0f}%)".format(model_data[0].get('reinvestment_pct', 50) * 100)
    html += "</td>"
    for year in years:
        val = model_data[year]['reinvestment_amount']
        css = 'positive' if val > 0 else ''
        html += f"<td class='{css}'>${val/1e6:.1f}M</td>"
    html += "</tr>"

    # Row: Distribution
    html += "<tr><td class='label'>Cash Distribution</td>"
    for year in years:
        val = model_data[year]['cash_distributions']
        css = 'positive' if val > 0 else ''
        html += f"<td class='{css}'>${val/1e6:.1f}M</td>"
    html += "</tr>"

    # Row: Cumulative AUM
    html += "<tr><td class='label'><strong>Cumulative AUM</strong></td>"
    for year in years:
        val = model_data[year]['cumulative_aum']
        html += f"<td><strong>${val/1e6:.0f}M</strong></td>"
    html += "</tr>"

    html += """
    </table>
    </div>
        </div>
    </body>
    </html>
    """
    return html


def generate_nav_tabs(current_page):
    """Generate navigation tabs for all pages."""
    tabs = [
        ('financial_model.html', 'Base Case'),
        ('financial_model_timeline.html', 'Timeline'),
        ('financial_model_optimized.html', 'Optimized'),
        ('financial_model_comparison.html', 'Comparison'),
        ('financial_model_sensitivity.html', 'Sensitivity'),
    ]

    nav_html = """
    <style>
        .nav-tabs {
            display: flex;
            border-bottom: 2px solid #4CAF50;
            margin-bottom: 20px;
            background-color: #f9f9f9;
        }
        .nav-tab {
            padding: 12px 20px;
            text-decoration: none;
            color: #333;
            border-bottom: 3px solid transparent;
            cursor: pointer;
            font-weight: 500;
            transition: all 0.3s ease;
        }
        .nav-tab:hover {
            background-color: #f0f0f0;
            color: #4CAF50;
        }
        .nav-tab.active {
            color: #4CAF50;
            border-bottom: 3px solid #4CAF50;
        }
    </style>
    <div class="nav-tabs">
    """

    for page_url, page_name in tabs:
        active_class = 'active' if current_page == page_url else ''
        nav_html += f'<a href="{page_url}" class="nav-tab {active_class}">{page_name}</a>'

    nav_html += '</div>'
    return nav_html


def generate_html_output(model_data, assets, exit_data, shareholder_returns, irr_data, params):
    """Generate HTML output of the financial model."""

    # Build cashflow table first
    rows = []
    rows.append("<tr><th>Year</th><th>New Deployment</th><th>Asset Returns</th><th>Debt Drawn</th><th>Mgmt Fee</th><th>Debt Interest</th><th>Available Cash</th><th>Reinvestment</th><th>Distribution</th><th>Cumulative AUM</th></tr>")

    for year in sorted(model_data.keys()):
        data = model_data[year]
        rows.append(f"""
            <tr>
                <td class="label">{year}</td>
                <td class="number">${data['new_deployment']:,.0f}</td>
                <td class="number">${data['asset_returns_gross']:,.0f}</td>
                <td class="number">${data['debt_drawn']:,.0f}</td>
                <td class="number">${data['mgmt_fee']:,.0f}</td>
                <td class="number">${data['debt_interest']:,.0f}</td>
                <td class="number">${data['available_cash']:,.0f}</td>
                <td class="number">${data['reinvestment_amount']:,.0f}</td>
                <td class="number">${data['cash_distributions']:,.0f}</td>
                <td class="number">${data['cumulative_aum']:,.0f}</td>
            </tr>
        """)

    cashflow_table = f"<table>{''.join(rows)}</table>"

    # Format parameters for display
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    exit_year = exit_data['exit_year']

    fee_income = shareholder_returns['cumulative_fee_income']
    upside_income = shareholder_returns['exit_upside_income']
    total_cash = shareholder_returns['total_cash_to_shareholder']
    shareholder_pct = shareholder_returns['shareholder_pct_of_mgmt'] * 100
    upside_pct = shareholder_returns['shareholder_upside_pct'] * 100

    nav_tabs = generate_nav_tabs('financial_model.html')

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Music Royalty Fund Financial Model</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
            h1, h2 {{ color: #333; }}
            table {{ border-collapse: collapse; width: 100%; margin: 20px 0; background-color: white; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: right; }}
            th {{ background-color: #4CAF50; color: white; text-align: left; }}
            td.label {{ text-align: left; font-weight: bold; }}
            .container {{ max-width: 1200px; margin: 0 auto; background-color: white; padding: 20px; border-radius: 5px; }}
            .summary-box {{ background-color: #e8f5e9; padding: 15px; margin: 20px 0; border-radius: 5px; }}
            .param-box {{ background-color: #fff3cd; padding: 10px; margin: 10px 0; border-radius: 5px; font-size: 0.9em; }}
            .number {{ text-align: right; }}
        </style>
    </head>
    <body>
        <div class="container">
            {nav_tabs}
            <h1>Music Royalty Fund Financial Model</h1>
            <p>Generated: {timestamp}</p>
            <p style="margin-top: 15px;">
                <a href="financial_model.xlsx" style="display: inline-block; padding: 10px 20px; background-color: #4CAF50; color: white; text-decoration: none; border-radius: 5px; font-weight: bold;">
                    ⬇️ Download Excel
                </a>
            </p>

            <h2>Model Parameters</h2>
            <div class="param-box">
                <strong>Fund:</strong> ${params.total_equity:,.0f} deployed over {params.deployment_years} years<br>
                <strong>Average Deal Size:</strong> ${params.avg_deal_size:,.0f}<br>
                <strong>Purchase Multiple:</strong> {params.purchase_multiple}x revenues<br>
                <strong>Annual Asset Return:</strong> {params.annual_asset_return:.1%}<br>
                <strong>Hold Period:</strong> {params.hold_period_years} years<br>
                <br>
                <strong>Deal Costs:</strong> Data {params.deal_cost_data_work:.1%} | Valuation {params.deal_cost_valuation:.1%} | Legal {params.deal_cost_legal:.1%} | Broker {params.deal_cost_broker:.1%}<br>
                <br>
                <strong>Management Co:</strong> 2% AUM fee | 20% upside | Owned 33.33% by UK company<br>
                <strong>Shareholder:</strong> 44% of UK company ({shareholder_pct:.2f}% of mgmt co, {upside_pct:.2f}% of upside)<br>
                <br>
                <strong>Debt:</strong> ${params.debt_amount:,.0f} at {params.debt_interest_rate:.1%} interest, raised mid-year 3<br>
                <strong>Debt Arrangement Fee:</strong> {params.debt_arrangement_fee:.1%}<br>
                <strong>Reinvestment Rate:</strong> {params.reinvestment_rate:.0%}<br>
                <strong>Exit Multiple:</strong> {params.exit_multiple_low}x - {params.exit_multiple_high}x<br>
            </div>

            <h2>Year-by-Year Cashflow Model</h2>
            {cashflow_table}

            <h2>Returns & IRR Analysis</h2>
            <div class="summary-box">
                <p><strong>Investor IRR (80% of upside):</strong> {irr_data['investor_irr']:.1f}%</p>
                <p style="font-weight: bold; color: {'green' if irr_data['investor_irr'] >= 20 else 'red'};">{'✓ MEETS 20% Target' if irr_data['investor_irr'] >= 20 else '✗ BELOW 20% Target'}</p>
            </div>

            <h2>Shareholder Income Breakdown</h2>
            <div class="summary-box">
                <p><strong>Fee Dividend Income ({shareholder_pct:.2f}% of mgmt fees):</strong> ${fee_income:,.0f}</p>
                <p><strong>Exit Upside ({upside_pct:.2f}% of net exit value):</strong> ${upside_income:,.0f}</p>
                <hr>
                <p><strong>TOTAL CASH TO SHAREHOLDER:</strong> ${total_cash:,.0f}</p>
            </div>

            <h2>Exit Summary (Year {exit_year})</h2>
            <div class="summary-box">
                <p><strong>Exit Value (at {params.exit_multiple_low}x):</strong> ${exit_data['exit_value']:,.0f}</p>
                <p><strong>Exit Costs ({params.exit_costs_pct:.1%}):</strong> ${exit_data['exit_costs']:,.0f}</p>
                <p><strong>Net Proceeds to Investors:</strong> ${exit_data['net_proceeds']:,.0f}</p>
            </div>
        </div>
    </body>
    </html>
    """

    return html


# ============================================================================
# RUN MODEL
# ============================================================================

def run_sensitivity_analysis(base_params):
    """Run sensitivity analysis on key IRR drivers."""
    sensitivity_results = {
        'exit_multiple': {},
        'asset_return': {},
        'mgmt_fee': {},
        'debt_interest': {},
        'reinvestment_rate': {}
    }

    # Test exit multiple (14x to 18x)
    for multiple in [14, 15, 16, 17, 18]:
        params = ModelParams()
        params.exit_multiple_low = multiple
        params.exit_multiple_high = multiple
        model_data, assets, _, _ = build_financial_model(params)
        exit_data = calculate_exit_value(model_data, assets, params, 0, 0)
        irr_data = calculate_irr(model_data, exit_data, params)
        sensitivity_results['exit_multiple'][multiple] = irr_data['investor_irr']

    # Test asset return (8% to 15%)
    for asset_ret in [0.08, 0.10, 0.12, 0.14, 0.15]:
        params = ModelParams()
        params.annual_asset_return = asset_ret
        model_data, assets, _, _ = build_financial_model(params)
        exit_data = calculate_exit_value(model_data, assets, params, 0, 0)
        irr_data = calculate_irr(model_data, exit_data, params)
        sensitivity_results['asset_return'][f'{asset_ret:.0%}'] = irr_data['investor_irr']

    # Test management fee (1% to 3%)
    for fee in [0.01, 0.015, 0.02, 0.025, 0.03]:
        params = ModelParams()
        params.mgmt_fee_rate = fee
        model_data, assets, _, _ = build_financial_model(params)
        exit_data = calculate_exit_value(model_data, assets, params, 0, 0)
        irr_data = calculate_irr(model_data, exit_data, params)
        sensitivity_results['mgmt_fee'][f'{fee:.1%}'] = irr_data['investor_irr']

    # Test debt interest rate (5% to 9%)
    for debt_rate in [0.05, 0.06, 0.07, 0.08, 0.09]:
        params = ModelParams()
        params.debt_interest_rate = debt_rate
        model_data, assets, _, _ = build_financial_model(params)
        exit_data = calculate_exit_value(model_data, assets, params, 0, 0)
        irr_data = calculate_irr(model_data, exit_data, params)
        sensitivity_results['debt_interest'][f'{debt_rate:.0%}'] = irr_data['investor_irr']

    # Test reinvestment rate (0% to 100%)
    for reinvest in [0.0, 0.25, 0.50, 0.75, 1.0]:
        params = ModelParams()
        params.reinvestment_rate = reinvest
        model_data, assets, _, _ = build_financial_model(params)
        exit_data = calculate_exit_value(model_data, assets, params, 0, 0)
        irr_data = calculate_irr(model_data, exit_data, params)
        sensitivity_results['reinvestment_rate'][f'{reinvest:.0%}'] = irr_data['investor_irr']

    return sensitivity_results


def generate_sensitivity_html(sensitivity_results):
    """Generate HTML sensitivity table."""
    nav_tabs = generate_nav_tabs('financial_model_sensitivity.html')
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>IRR Sensitivity Analysis</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }
            h1, h2 { color: #333; }
            table { border-collapse: collapse; width: 100%; margin: 20px 0; background-color: white; }
            th, td { border: 1px solid #ddd; padding: 12px; text-align: center; }
            th { background-color: #4CAF50; color: white; }
            td.label { text-align: left; font-weight: bold; background-color: #f9f9f9; }
            .container { max-width: 1000px; margin: 0 auto; background-color: white; padding: 20px; border-radius: 5px; }
            .target { color: white; background-color: #4CAF50; font-weight: bold; }
            .below-target { color: white; background-color: #ff9800; }
            .above-target { color: white; background-color: #2196F3; }
        </style>
    </head>
    <body>
        <div class="container">
            """ + nav_tabs + """
            <h1>IRR Sensitivity Analysis — Key Levers</h1>
            <p style="margin-bottom: 15px;">
                <a href="financial_model.xlsx" style="display: inline-block; padding: 10px 20px; background-color: #4CAF50; color: white; text-decoration: none; border-radius: 5px; font-weight: bold;">
                    ⬇️ Download Excel
                </a>
            </p>
            <p><strong>Target IRR:</strong> 20%</p>
    """

    # Exit Multiple Table
    html += "<h2>Exit Multiple Impact</h2>"
    html += "<table><tr><th>Exit Multiple</th><th>Investor IRR</th><th>vs 20% Target</th></tr>"
    for multiple in sorted(sensitivity_results['exit_multiple'].keys()):
        irr = sensitivity_results['exit_multiple'][multiple]
        status_class = 'target' if abs(irr - 20) < 0.5 else ('above-target' if irr >= 20 else 'below-target')
        gap = irr - 20
        html += f"<tr><td class='label'>{multiple}x</td><td class='{status_class}'>{irr:.1f}%</td><td>{gap:+.1f}%</td></tr>"
    html += "</table>"

    # Asset Return Table
    html += "<h2>Asset Return Impact</h2>"
    html += "<table><tr><th>Annual Return</th><th>Investor IRR</th><th>vs 20% Target</th></tr>"
    for ret_str in sorted(sensitivity_results['asset_return'].keys(), key=lambda x: float(x.rstrip('%'))/100):
        irr = sensitivity_results['asset_return'][ret_str]
        status_class = 'target' if abs(irr - 20) < 0.5 else ('above-target' if irr >= 20 else 'below-target')
        gap = irr - 20
        html += f"<tr><td class='label'>{ret_str}</td><td class='{status_class}'>{irr:.1f}%</td><td>{gap:+.1f}%</td></tr>"
    html += "</table>"

    # Management Fee Table
    html += "<h2>Management Fee Impact</h2>"
    html += "<table><tr><th>Annual Fee</th><th>Investor IRR</th><th>vs 20% Target</th></tr>"
    for fee_str in sorted(sensitivity_results['mgmt_fee'].keys(), key=lambda x: float(x.rstrip('%'))/100):
        irr = sensitivity_results['mgmt_fee'][fee_str]
        status_class = 'target' if abs(irr - 20) < 0.5 else ('above-target' if irr >= 20 else 'below-target')
        gap = irr - 20
        html += f"<tr><td class='label'>{fee_str}</td><td class='{status_class}'>{irr:.1f}%</td><td>{gap:+.1f}%</td></tr>"
    html += "</table>"

    # Debt Interest Table
    html += "<h2>Debt Interest Rate Impact</h2>"
    html += "<table><tr><th>Interest Rate</th><th>Investor IRR</th><th>vs 20% Target</th></tr>"
    for rate_str in sorted(sensitivity_results['debt_interest'].keys(), key=lambda x: float(x.rstrip('%'))/100):
        irr = sensitivity_results['debt_interest'][rate_str]
        status_class = 'target' if abs(irr - 20) < 0.5 else ('above-target' if irr >= 20 else 'below-target')
        gap = irr - 20
        html += f"<tr><td class='label'>{rate_str}</td><td class='{status_class}'>{irr:.1f}%</td><td>{gap:+.1f}%</td></tr>"
    html += "</table>"

    # Reinvestment Rate Table
    html += "<h2>Reinvestment Rate Impact</h2>"
    html += "<table><tr><th>Reinvestment Rate</th><th>Investor IRR</th><th>vs 20% Target</th></tr>"
    for reinvest_str in sorted(sensitivity_results['reinvestment_rate'].keys(), key=lambda x: float(x.rstrip('%'))/100):
        irr = sensitivity_results['reinvestment_rate'][reinvest_str]
        status_class = 'target' if abs(irr - 20) < 0.5 else ('above-target' if irr >= 20 else 'below-target')
        gap = irr - 20
        html += f"<tr><td class='label'>{reinvest_str}</td><td class='{status_class}'>{irr:.1f}%</td><td>{gap:+.1f}%</td></tr>"
    html += "</table>"

    html += """
        </div>
    </body>
    </html>
    """
    return html


def generate_comparison_html(base_case, optimized_case, base_params, opt_params):
    """Generate HTML comparing base case vs optimized case."""
    nav_tabs = generate_nav_tabs('financial_model_comparison.html')
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Base Case vs Optimized Scenario</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
            h1, h2 {{ color: #333; }}
            table {{ border-collapse: collapse; width: 100%; margin: 20px 0; background-color: white; }}
            th, td {{ border: 1px solid #ddd; padding: 12px; text-align: right; }}
            th {{ background-color: #4CAF50; color: white; text-align: left; }}
            td.label {{ text-align: left; font-weight: bold; }}
            .container {{ max-width: 1200px; margin: 0 auto; background-color: white; padding: 20px; border-radius: 5px; }}
            .summary-box {{ background-color: #e8f5e9; padding: 15px; margin: 20px 0; border-radius: 5px; }}
            .param-box {{ background-color: #fff3cd; padding: 10px; margin: 10px 0; border-radius: 5px; font-size: 0.9em; }}
            .number {{ text-align: right; }}
            .highlight {{ background-color: #fff59d; font-weight: bold; }}
            .success {{ background-color: #c8e6c9; font-weight: bold; color: green; }}
        </style>
    </head>
    <body>
        <div class="container">
            {nav_tabs}
            <h1>Base Case vs Optimized Scenario</h1>
            <p style="margin-bottom: 20px;">
                <a href="financial_model.xlsx" style="display: inline-block; padding: 10px 20px; background-color: #4CAF50; color: white; text-decoration: none; border-radius: 5px; font-weight: bold;">
                    ⬇️ Download Excel
                </a>
            </p>

            <h2>Parameter Changes</h2>
            <table>
                <tr><th>Parameter</th><th>Base Case</th><th>Optimized</th><th>Change</th></tr>
                <tr><td class="label">Exit Multiple</td><td>{base_params.exit_multiple_low}x</td><td>{opt_params.exit_multiple_low}x</td><td>{opt_params.exit_multiple_low - base_params.exit_multiple_low}x</td></tr>
                <tr><td class="label">Management Fee</td><td>{base_params.mgmt_fee_rate:.1%}</td><td>{opt_params.mgmt_fee_rate:.1%}</td><td>{opt_params.mgmt_fee_rate - base_params.mgmt_fee_rate:.1%}</td></tr>
                <tr><td class="label">Reinvestment Rate</td><td>{base_params.reinvestment_rate:.0%}</td><td>{opt_params.reinvestment_rate:.0%}</td><td>{(opt_params.reinvestment_rate - base_params.reinvestment_rate):.0%}</td></tr>
                <tr><td class="label">Debt Interest Rate</td><td>{base_params.debt_interest_rate:.1%}</td><td>{opt_params.debt_interest_rate:.1%}</td><td>{opt_params.debt_interest_rate - base_params.debt_interest_rate:.1%}</td></tr>
            </table>

            <h2>Results Comparison</h2>
            <table>
                <tr><th>Metric</th><th>Base Case</th><th>Optimized</th><th>Improvement</th></tr>
                <tr><td class="label">Investor IRR</td>
                    <td class="highlight">{base_case['investor_irr']:.1f}%</td>
                    <td class="{'success' if optimized_case['investor_irr'] >= 20 else 'highlight'}">{optimized_case['investor_irr']:.1f}%</td>
                    <td class="{'success' if optimized_case['investor_irr'] >= 20 else ''}">{optimized_case['investor_irr'] - base_case['investor_irr']:+.1f}%</td>
                </tr>
                <tr><td class="label">Cumulative AUM</td><td>{base_case['aum']:,.0f}</td><td>{optimized_case['aum']:,.0f}</td><td>{optimized_case['aum'] - base_case['aum']:,.0f}</td></tr>
                <tr><td class="label">Cumulative Fee Income</td><td>${base_case['fee_income']:,.0f}</td><td>${optimized_case['fee_income']:,.0f}</td><td>${optimized_case['fee_income'] - base_case['fee_income']:,.0f}</td></tr>
                <tr><td class="label">Exit Upside Income</td><td>${base_case['upside_income']:,.0f}</td><td>${optimized_case['upside_income']:,.0f}</td><td>${optimized_case['upside_income'] - base_case['upside_income']:,.0f}</td></tr>
                <tr><td class="label"><strong>Total Cash to Shareholder</strong></td>
                    <td class="highlight">${base_case['total_cash']:,.0f}</td>
                    <td class="{'success' if optimized_case['investor_irr'] >= 20 else 'highlight'}">${optimized_case['total_cash']:,.0f}</td>
                    <td class="{'success' if optimized_case['investor_irr'] >= 20 else ''}">${optimized_case['total_cash'] - base_case['total_cash']:+,.0f}</td>
                </tr>
            </table>

            <div class="summary-box">
                <h3>Key Insights</h3>
                <ul>
                    <li><strong>Exit Multiple:</strong> Increased from {base_params.exit_multiple_low}x to {opt_params.exit_multiple_low}x — realistic within {base_params.exit_multiple_low}-{base_params.exit_multiple_high}x range</li>
                    <li><strong>Management Fee:</strong> Reduced from {base_params.mgmt_fee_rate:.1%} to {opt_params.mgmt_fee_rate:.1%} — achievable through negotiation or tiered structure</li>
                    <li><strong>Reinvestment Strategy:</strong> Reduced from {base_params.reinvestment_rate:.0%} to {opt_params.reinvestment_rate:.0%} — distribute more, reinvest less</li>
                    <li><strong>Debt Cost:</strong> Improved from {base_params.debt_interest_rate:.1%} to {opt_params.debt_interest_rate:.1%} — better market conditions or structure</li>
                </ul>
                <p style="font-weight: bold; color: green; margin-top: 15px;">✓ Optimized scenario achieves {optimized_case['investor_irr']:.1f}% IRR, {'EXCEEDING' if optimized_case['investor_irr'] >= 20 else 'approaching'} the 20% target.</p>
            </div>
        </div>
    </body>
    </html>
    """
    return html


def export_to_excel(model_data_base, shareholder_base, irr_base, params_base,
                     model_data_opt, shareholder_opt, irr_opt, params_opt,
                     sensitivity_results):
    """Export financial model to Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Base Case"

    # Style definitions
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    number_format = "#,##0"
    currency_format = '$#,##0'
    pct_format = '0.0%'

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Base Case Year-by-Year
    years = sorted(model_data_base.keys())

    # Headers
    ws['A1'] = "Base Case — Year-by-Year Cashflow Model"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:J1')

    row = 3
    headers = ['Year', 'New Deployment', 'Asset Returns', 'Debt Drawn', 'Mgmt Fee',
               'Debt Interest', 'Available Cash', 'Reinvestment', 'Distribution', 'Cumulative AUM']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    row = 4
    for year in years:
        data = model_data_base[year]
        row_data = [
            year,
            data['new_deployment'],
            data['asset_returns_gross'],
            data['debt_drawn'],
            data['mgmt_fee'],
            data['debt_interest'],
            data['available_cash'],
            data['reinvestment_amount'],
            data['cash_distributions'],
            data['cumulative_aum']
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            if col == 1:
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.number_format = currency_format
                cell.alignment = Alignment(horizontal='right')
            cell.border = thin_border
        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col].width = 16

    # Sheet 2: Summary
    ws_summary = wb.create_sheet("Summary")

    ws_summary['A1'] = "Financial Model Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)

    row = 3
    ws_summary[f'A{row}'] = "Base Case"
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)

    row = 4
    summary_data = [
        ('Investor IRR', irr_base['investor_irr'], pct_format),
        ('Cumulative AUM', sum([model_data_base[y]['cumulative_aum'] for y in years]) / len(years), currency_format),
        ('Cumulative Fee Income', shareholder_base['cumulative_fee_income'], currency_format),
        ('Exit Upside Income', shareholder_base['exit_upside_income'], currency_format),
        ('Total Cash to Shareholder', shareholder_base['total_cash_to_shareholder'], currency_format),
    ]

    for label, value, fmt in summary_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        ws_summary[f'B{row}'].number_format = fmt
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1

    row += 1
    ws_summary[f'A{row}'] = "Optimized Case"
    ws_summary[f'A{row}'].font = Font(bold=True, size=12, color="00AA00")

    row += 1
    summary_opt_data = [
        ('Investor IRR', irr_opt['investor_irr'], pct_format),
        ('Total Cash to Shareholder', shareholder_opt['total_cash_to_shareholder'], currency_format),
    ]

    for label, value, fmt in summary_opt_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        ws_summary[f'B{row}'].number_format = fmt
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1

    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20

    # Sheet 3: Sensitivity
    ws_sens = wb.create_sheet("Sensitivity")

    ws_sens['A1'] = "IRR Sensitivity Analysis"
    ws_sens['A1'].font = Font(bold=True, size=14)
    ws_sens.merge_cells('A1:C1')

    row = 3

    # Exit Multiple Sensitivity
    ws_sens[f'A{row}'] = "Exit Multiple"
    ws_sens[f'A{row}'].font = Font(bold=True, color="FFFFFF")
    ws_sens[f'A{row}'].fill = header_fill
    ws_sens[f'B{row}'] = "Investor IRR"
    ws_sens[f'B{row}'].font = Font(bold=True, color="FFFFFF")
    ws_sens[f'B{row}'].fill = header_fill
    row += 1

    for multiple in sorted(sensitivity_results['exit_multiple'].keys()):
        ws_sens[f'A{row}'] = f"{multiple}x"
        ws_sens[f'B{row}'] = sensitivity_results['exit_multiple'][multiple] / 100
        ws_sens[f'B{row}'].number_format = pct_format
        row += 1

    ws_sens.column_dimensions['A'].width = 15
    ws_sens.column_dimensions['B'].width = 15

    # Save workbook
    output_path = Path("/Users/andrewgoodwin/financial_model.xlsx")
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    # BASE CASE
    params_base = ModelParams()
    model_data_base, assets_base, aum_base, reinvest_base = build_financial_model(params_base)
    exit_data_base = calculate_exit_value(model_data_base, assets_base, params_base, aum_base, reinvest_base)
    shareholder_base = calculate_shareholder_returns(model_data_base, exit_data_base, params_base)
    irr_base = calculate_irr(model_data_base, exit_data_base, params_base)

    # Generate base model HTML
    html_base = generate_html_output(model_data_base, assets_base, exit_data_base, shareholder_base, irr_base, params_base)
    output_base = Path("/Users/andrewgoodwin/financial_model.html")
    output_base.write_text(html_base)
    print(f"✓ Base model generated: {output_base}")

    # Generate timeline view
    html_timeline = generate_timeline_html(model_data_base, params_base)
    output_timeline = Path("/Users/andrewgoodwin/financial_model_timeline.html")
    output_timeline.write_text(html_timeline)
    print(f"✓ Timeline view generated: {output_timeline}")

    # OPTIMIZED CASE
    params_opt = ModelParams()
    params_opt.exit_multiple_low = 16  # Increase from 14x to 16x
    params_opt.exit_multiple_high = 16
    params_opt.mgmt_fee_rate = 0.015  # Reduce from 2.0% to 1.5%
    params_opt.reinvestment_rate = 0.25  # Reduce from 50% to 25% (distribute more)
    params_opt.debt_interest_rate = 0.065  # Reduce from 7% to 6.5%

    model_data_opt, assets_opt, aum_opt, reinvest_opt = build_financial_model(params_opt)
    exit_data_opt = calculate_exit_value(model_data_opt, assets_opt, params_opt, aum_opt, reinvest_opt)
    shareholder_opt = calculate_shareholder_returns(model_data_opt, exit_data_opt, params_opt)
    irr_opt = calculate_irr(model_data_opt, exit_data_opt, params_opt)

    # Generate optimized model HTML
    html_opt = generate_html_output(model_data_opt, assets_opt, exit_data_opt, shareholder_opt, irr_opt, params_opt)
    output_opt = Path("/Users/andrewgoodwin/financial_model_optimized.html")
    output_opt.write_text(html_opt)
    print(f"✓ Optimized model generated: {output_opt}")

    # Generate comparison
    base_case_data = {
        'investor_irr': irr_base['investor_irr'],
        'aum': aum_base,
        'fee_income': shareholder_base['cumulative_fee_income'],
        'upside_income': shareholder_base['exit_upside_income'],
        'total_cash': shareholder_base['total_cash_to_shareholder'],
    }

    opt_case_data = {
        'investor_irr': irr_opt['investor_irr'],
        'aum': aum_opt,
        'fee_income': shareholder_opt['cumulative_fee_income'],
        'upside_income': shareholder_opt['exit_upside_income'],
        'total_cash': shareholder_opt['total_cash_to_shareholder'],
    }

    html_comparison = generate_comparison_html(base_case_data, opt_case_data, params_base, params_opt)
    output_comparison = Path("/Users/andrewgoodwin/financial_model_comparison.html")
    output_comparison.write_text(html_comparison)
    print(f"✓ Comparison generated: {output_comparison}")

    # Generate sensitivity analysis
    print("\n📊 Running sensitivity analysis...")
    sensitivity_results = run_sensitivity_analysis(params_base)
    sensitivity_html = generate_sensitivity_html(sensitivity_results)
    sensitivity_path = Path("/Users/andrewgoodwin/financial_model_sensitivity.html")
    sensitivity_path.write_text(sensitivity_html)
    print(f"✓ Sensitivity analysis generated: {sensitivity_path}")

    # Export to Excel
    excel_path = export_to_excel(model_data_base, shareholder_base, irr_base, params_base,
                                  model_data_opt, shareholder_opt, irr_opt, params_opt,
                                  sensitivity_results)
    print(f"✓ Excel export generated: {excel_path}")

    print(f"\n📈 Summary:")
    print(f"\n  BASE CASE:")
    print(f"    Investor IRR: {irr_base['investor_irr']:.1f}% (gap: {irr_base['investor_irr']-20:.1f}%)")
    print(f"    Shareholder receives: ${shareholder_base['total_cash_to_shareholder']:,.0f}")
    print(f"\n  OPTIMIZED CASE:")
    print(f"    Investor IRR: {irr_opt['investor_irr']:.1f}% ✓ HITS TARGET" if irr_opt['investor_irr'] >= 20 else f"    Investor IRR: {irr_opt['investor_irr']:.1f}% (gap: {irr_opt['investor_irr']-20:.1f}%)")
    print(f"    Shareholder receives: ${shareholder_opt['total_cash_to_shareholder']:,.0f}")
    print(f"    Improvement: +${shareholder_opt['total_cash_to_shareholder'] - shareholder_base['total_cash_to_shareholder']:,.0f}")
